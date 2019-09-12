#!/usr/bin/env python3

import openpyxl

wb1 = openpyxl.load_workbook('AllandEverything.xlsx')
ws1 = wb1.active
ws2 = wb1.create_sheet("unique")

values = []
for i in range(2,ws1.max_row+1):
    if ws1.cell(row=i,column=1).value in values:
        pass
    else:
        values.append(ws1.cell(row=i,column=1).value)


for value in values:
    ws2.append([value])

wb1.save('AllandEverything.xlsx')
